import datetime
import sys
import os.path
import openpyxl
import os
import keyboard
from termcolor import colored, cprint
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
# 1. khai bao va khoi chay  web driver
options = webdriver.EdgeOptions()
options.page_load_strategy = 'normal'
options.add_argument("--headless")
service = EdgeService(executable_path='.\msedgedriver.exe')
# 2. thuc hien dang nhap  va dieu huong vao trang can lay du lieu
Usr = "kv2zoom@gmail.com"
pas = "27bnguyenthanhhan"
word = ''
driver = webdriver.Edge(service=service, options=options)
driver.get("https://eticket.danang.gov.vn/#/login")
driver.find_element(By.ID, 'username').send_keys(Usr)
driver.find_element(By.ID, 'password').send_keys(pas)
driver.find_element(
    By.XPATH, "/html/body/div/div/div/form/div[2]/button").click()
driver.implicitly_wait(1)
driver.find_element(By.PARTIAL_LINK_TEXT, 'check-ins').click()
# 3. lay ten cot ( header th) va luu vao mang
header = []
th_s = driver.find_elements(By.TAG_NAME, 'th')
for cot in th_s:
    header.append(cot.text)
header.append('ĐĂNG KÍ?')
# print(len(header))
# 4. check va mo file excel luu tru
file_name = datetime.datetime.now().strftime('%d_%m_%Y')+'.xlsx'
print(file_name)
PATH = './log/'+file_name
if os.path.isfile(PATH):
    # print('da ton tai')
    wb = openpyxl.load_workbook(filename=PATH)
    ws = wb.active
else:
    # print('chau ton tai')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NHAT KY NGAY"
    ws.append(header[1:])
    # cells=ws['A1':'O1']
    #  cells.font = openpyxl.styles.Font(bold="true",size=16)
    # wb.save(PATH)
##lay thoi diem gan nhat checkin 

# print(last_time_checkin)
## 5. doc du lieu va dua vao excel
word = input('Moi Ban Nhap C de thuc thi: ')
while(word.lower() != 'q'):
    if word=='c':
        print('DANG THUC HIEN QUET DU LIEU TREN HE THONG ETICKET.DANANG.GOV.VN')
        driver.refresh()
        time.sleep(1)
        row_data=driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div/main/div[2]/div/div[2]/div/table/tbody/tr[1]')
        td_s=row_data.find_elements(By.TAG_NAME,'td')
        cmnd_log_in=td_s[5].text
        HoTen=td_s[2].text
        print('HO va Ten nhan su dang checkin : '+HoTen)
        print('CMND/CCCD ĐANG THUC HIEN CHECK IN TAI CONG:'+cmnd_log_in)
        # driver.close()
        #mo file excel csdl nguoi dung de tim kiem
        # cur_dic=os.getcwd()
        wb2 = openpyxl.load_workbook('./data_usr.xlsx')
        sheet = wb2['data']
        DK=0
        max_row=sheet.max_row
        for i in range(2,max_row):
            DK=0
            ds_CMND=str(sheet.cell(row=i, column=9).value).strip().lower()
            if cmnd_log_in in ds_CMND:
                DK=1
                break
        
        if DK==1 :
            # text = colored('DA DANG KI TRONG DANH SACH', 'blue', attrs=['reverse', 'blink'])
            cprint("DA DANG KI TRONG DANH SACH", 'blue', attrs=['bold'], file=sys.stderr)
            register='ĐÃ ĐĂNG KÍ'
        else:
            cprint("CHUA DANG KI TRONG DANH SACH", 'red', attrs=['bold'], file=sys.stderr)
            register='CHƯA ĐĂNG KÍ'
        #sau khi checkin thi ghi du lieu vao log excel
        last_time_checkin = ws.cell(row=ws.max_row , column=12).value
        if td_s[12].text != last_time_checkin:
            data_to_write=[]
            for cot in td_s:
                data_to_write.append(cot.text)
            data_to_write.append(register)
            ws.append(data_to_write[1:])
        # print(data_to_write)c

    # word=input('Nhan Q de thoat , nhan C de tiep tuc :')
    print('Nhan Q de thoat , nhan phim bat ki  de tiep tuc :')
    if keyboard.read_key() == "q":
        driver.quit()
        wb.save(PATH)
        break
