import datetime
import sys
import os.path
import openpyxl
import os
import keyboard
import time
from termcolor import colored, cprint
from datetime import datetime
file_name = datetime.now().strftime('%d_%m_%Y')+'.xlsx'
PATH = './log/'+file_name


def open_wb():

    if os.path.isfile(PATH):
        wb = openpyxl.load_workbook(filename=PATH)
        # return wb
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('NHAT KY NGAY')
        wb.save(PATH)
        # ws.title = "NHAT KY NGAY"
        # return 0
    return wb

def openws(header, file_name):
    # file_name = datetime.datetime.now().strftime('%d_%m_%Y')+'.xlsx'
    print(file_name)
    PATH = './log/'+file_name
    if os.path.isfile(PATH):
        # print('da ton tai')
        wb = openpyxl.load_workbook(filename=PATH)
        ws = wb.active
    else:
        # print('chau ton tai')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NHAT KY NGAY"
        ws.append(header[1:])
    return wb


def check_open_filelog(PATH):
    if os.path.isfile(PATH):
        # print('da ton tai')
        try:
            wb = openpyxl.load_workbook(filename=PATH)
        except PermissionError:
            print('file log đang mở , vui lòng đóng file log để thực hiện tiếp')
            # dem_clear=dem_clear+1
            # time.sleep(3)
            # continue
            return 0
    else:
        print('file chưa tồn tại!!!')
    return 1


def check_save_filelog(wb, file_name):
    PATH = './log/'+file_name
    if os.path.isfile(PATH):
        # print('da ton tai')
        try:
            wb.save(PATH)
        except PermissionError:
            print('file log đang mở , vui lòng đóng file log để thực hiện tiếp')
            # dem_clear=dem_clear+1
            # time.sleep(3)
            # continue
            return 0
    else:
        print('file log chưa tồn tại!!!')
    return 1


def compare_ngay(ngay_het_han, ngay_hien_tai):
    year_het_han=datetime.strptime(ngay_het_han, "%d-%m-%Y").strftime("%Y")
    year_hien_tai=datetime.strptime(ngay_hien_tai, "%d-%m-%Y").strftime("%Y")
    if year_het_han >= year_hien_tai:
        ngay_het_han = int((datetime.strptime(ngay_het_han, "%d-%m-%Y")).strftime("%j"))+365*(int(year_het_han)-int(year_hien_tai))
        ngay_hien_tai = int((datetime.strptime(ngay_hien_tai, "%d-%m-%Y")).strftime("%j"))
        thoi_han=ngay_het_han-ngay_hien_tai
    else:
        thoi_han=0
    if thoi_han>0 :
        return 1 # con han
    else:
        return 0  # 'het HẠN'
