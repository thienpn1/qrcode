import openpyxl
import os

import keyboard
from selenium import webdriver
# from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver import Firefox 
from selenium.webdriver.firefox.service import Service
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
serv = Service('.\geckodriver.exe')
options = webdriver.FirefoxOptions()
options.page_load_strategy = 'normal'
options.binary_location=r'D:\Soft\FirefoxPortable\App\firefox64\firefox.exe'
options.add_argument("--headless")
# service = Service(executable_path='.\geckodriver.exe')
Usr = "kv2zoom@gmail.com"
pas = "27bnguyenthanhhan"
word = ''
driver = Firefox(service=serv, options=options)
driver.get("https://eticket.danang.gov.vn/#/login")
driver.find_element(By.ID,'username').send_keys(Usr)
driver.find_element(By.ID,'password').send_keys(pas)
driver.find_element(By.XPATH,"/html/body/div/div/div/form/div[2]/button").click()
driver.implicitly_wait(1)
driver.find_element(By.PARTIAL_LINK_TEXT,'check-ins').click()
word = input('Moi Ban Nhap C de thuc thi: ')
while(word.lower() != 'q'):
    
    if word=='c':
        driver.refresh()
        print('DANG THUC HIEN QUET DU LIEU TREN HE THONG ETICKET.DANANG.GOV.VN')
        
        time.sleep(1)
        row_data=driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div/main/div[2]/div/div[2]/div/table/tbody/tr[1]')
        td_s=row_data.find_elements(By.TAG_NAME,'td')
        cmnd_log_in=td_s[5].text
        HoTen=td_s[2].text
        print('HO va Ten nhan su dang checkin : '+HoTen)
        print('CMND/CCCD ĐANG THUC HIEN CHECK IN TAI CONG:'+cmnd_log_in)
        # driver.close()
        cur_dic=os.getcwd()
        wb = openpyxl.load_workbook('./data_usr.xlsx')
        sheet = wb['data']
        DK=0
        max_row=sheet.max_row
        for i in range(2,max_row):
            DK=0
            ds_CMND=str(sheet.cell(row=i, column=9).value).strip().lower()
            if cmnd_log_in in ds_CMND:
                DK=1
                break 
        if DK==1 :
            print('DA DANG KI TRONG DANH SACH')
           
        else:
            print('CHUA DANG KI TRONG DANH SACH')
    # word=input('Nhan Q de thoat , nhan C de tiep tuc :')
    print('Nhan Q de thoat , nhan phim bat ki  de tiep tuc :')
    if keyboard.read_key() == "q":
        driver.quit()
        break
        

